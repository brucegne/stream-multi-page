import streamlit as st
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")
st.write("It worked!")

workbook = load_workbook(filename="hello_world.xlsx")
st.write(workbook.sheetnames)


sheet = workbook.active
st.write(sheet)


st.write(sheet.title)
